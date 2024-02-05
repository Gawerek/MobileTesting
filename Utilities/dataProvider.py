import openpyxl

def get_data(sheetName):
    workbook = openpyxl.load_workbook(r"C:\Projekty_Python\WellmifyTests\Excel\testdata.xlsx") #@TODO zmiana na sciezke relatywna
    sheet = workbook[sheetName]
    mainList = []

    for row in range(2, sheet.max_row + 1):
        # Assuming that the fifth column in Excel contains the service names separated by semicolons
        service_names = sheet.cell(row=row, column=5).value
        service_list = service_names.split(';') if service_names else []
        mainList.append([
            sheet.cell(row=row, column=1).value,  # service_info
            sheet.cell(row=row, column=2).value,  # name
            sheet.cell(row=row, column=3).value,  # service_type
            sheet.cell(row=row, column=4).value,  # address
            service_list,                         # service_names
        ])
    return mainList